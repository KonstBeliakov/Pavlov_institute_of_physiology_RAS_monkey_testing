import json
import time
from tkinter import *

import mss
from PIL import Image
from PIL.ImageTk import PhotoImage

import pygame
import serial
import threading
from dateutil import parser
import datetime
from os import listdir
import pandas as pd
import openpyxl
import copy

from settings import settings, APPLICATION_RUNNING

pygame.init()
serial_available = False
ser = None

arduino_data = [0] * 8


def check_serial():
    global ser, serial_available
    try:
        if ser is not None:
            ser.close()
        ser = serial.Serial(settings['port'], settings['baudrate'])
        serial_available = True
    except Exception as e:
        print(e)
        serial_available = False
    print(f'serial_available: {serial_available}')
    return serial_available


def open_image(filename: str, image_size: int):
    return PhotoImage(Image.open(filename).resize((image_size, image_size)))


def entry_value_check(value, name, declension=1, min_value=None, max_value=None, value_type=int, may_be_empty=False):
    if not value:
        if may_be_empty:
            return None
        return f'{name} не указан{["", "о", "а"][declension]}'
    if value_type == 'list float':
        try:
            t = [float(i) for i in value.split()]
        except:
            return f'{name} долж{["ен", "но", "на"][declension]} быть списком вещественных чисел (через пробел)'
        else:
            if min_value is not None:
                for i in t:
                    if i < min_value:
                        return f'Все числа в списке "{name}" должны быть больше или равны {min_value}'
            if max_value is not None:
                for i in t:
                    if i > max_value:
                        return f'Все числа в списке "{name}" должны быть меньше или равны {max_value}'
            return None
    elif value_type == 'date':
        try:
            value = parser.parse(value).date()
        except:
            return f'Поле "{name}" должно быть датой (YYYY-MM-DD)'
    try:
        if value_type == int:
            int(value)
        elif value_type == float:
            float(value)
    except:
        d = {int: 'целым', float: 'вещественным'}
        return f'{name} долж{["ен", "но", "на"][declension]} быть {d[value_type]} числом'
    if min_value is not None and min_value > value_type(value):
        return f'{name} не может быть меньше {min_value}'
    if max_value is not None and max_value < value_type(value):
        return f'{name} не может быть больше {max_value}'


def right_answer():
    if settings['using_sound'] and settings['right_answer_sound']:
        play_sound(settings['right_answer_sound'])
    positive_reinforcement()


def wrong_answer():
    if settings['using_sound'] and settings['wrong_answer_sound']:
        play_sound(settings['wrong_answer_sound'])


def experiment_start():
    if settings['using_sound'] and settings['experiment_start_sound']:
        play_sound(settings['experiment_start_sound'])


def test_start():
    if settings['using_sound'] and settings['test_start_sound']:
        play_sound(settings['test_start_sound'])


def play_sound(sound):
    pygame.mixer.music.load(sound)
    pygame.mixer.music.play(0)


def send_data():
    x = sum([value * 2 ** i for i, value in enumerate(arduino_data)])
    ser.write(bytes([x]))


def positive_reinforcement():
    if serial_available:
        if settings['arduino_script'] == '24-07-18.ino':
            def f():
                ser.write(bytes([1]))
                time.sleep(settings['drink_delay'])
                ser.write(bytes([2]))

            t = threading.Thread(target=f)
            t.start()
        else:
            def f():
                arduino_data[0] = 1
                send_data()
                time.sleep(settings['drink_delay'])
                arduino_data[0] = 0
                send_data()

            t = threading.Thread(target=f)
            t.start()


def disable_anser_entry():
    if serial_available:
        if settings['arduino_script'] == '24-07-18.ino':
            ser.write(bytes([3]))
        else:
            def f():
                arduino_data[1] = 1
                send_data()
                time.sleep(settings['barrier_working_time'])
                arduino_data[1] = 0
                send_data()

            t = threading.Thread(target=f)
            t.start()


def anable_answer_entry():
    if serial_available:
        if settings['arduino_script'] == '24-07-18.ino':
            ser.write(bytes([4]))
        else:
            def f():
                arduino_data[2] = 1
                send_data()
                time.sleep(settings['barrier_working_time'])
                arduino_data[2] = 0
                send_data()

            t = threading.Thread(target=f)
            t.start()


def read_usb():
    while APPLICATION_RUNNING:
        if serial_available:
            print(f'received usb message: {ser.readline()}')
        else:
            print('serial is not available')
        time.sleep(.1)


def export_data(export_file=None, files=None):
    if export_file is None:
        export_file = f'datasets/dataset_{str(datetime.datetime.now()).split(".")[0].replace(":", "_")}.xlsx'
    if files is None:
        files = [i for i in listdir('data/') if i.endswith('.xlsx')]

    l = []

    for file in sorted(files):
        dataframe = openpyxl.load_workbook(file)
        dataframe1 = dataframe.active
        for row in range(1, dataframe1.max_row):
            l.append([col[row].value for col in dataframe1.iter_cols(3, dataframe1.max_column)] + [file])

    l = [[i + 1] + row for i, row in enumerate(l)]
    l = [['Номер эксперимента', 'Абсолютное время', 'Время на ответ', 'Ответ', 'Правильный ответ',
          'Источник информации']] + l

    data_frame = pd.DataFrame(l)
    data_frame.to_excel(export_file)


def load_settings(filename=None):
    if filename is None:
        filename = settings['settings_filename']
    print('Загружаем настройки...', end='')
    try:
        with open(filename, 'r') as settings_file:
            data = json.load(settings_file)
            for key in data:
                settings[key] = data[key]
    except Exception as err:
        print(f'\nПри загрузке настроек произошла ошибка: {err}')
    else:
        print('Успешно')


def save_settings(filename=None):
    if filename is None:
        filename = settings['settings_filename']
    settings1 = copy.deepcopy(settings)
    del settings1['selected_period_start']  # json can't save date
    del settings1['selected_period_end']
    print('Сохранение настроек...', end='')
    with open(filename, 'w') as settings_file:
        json.dump(settings1, settings_file)
    print('Успешно')


def auto_save(filename=None):
    t = time.perf_counter()
    while APPLICATION_RUNNING:
        if time.perf_counter() - t > settings['autosave_period'] * 60:
            t += settings['autosave_period'] * 60
            save_settings(filename)
        time.sleep(1)


def start_auto_saving(filename=None):
    t = threading.Thread(target=lambda: auto_save(filename))
    t.start()


def move_to_first_screen(window):
    with mss.mss() as sct:
        if len(sct.monitors) > 2:
            mon = sct.monitors[1]
            window.geometry(f'-{mon['width']}+0')


def paint_second_monitor():
    window = Tk()
    with mss.mss() as sct:
        if len(sct.monitors) > 2:
            mon = sct.monitors[2]
            window.geometry(f'{mon["width"]}x{mon["width"]}')
        canvas = Canvas(window, bg=settings['bg_color'], width=mon["width"], height=mon["width"])
        canvas.pack()
    window.overrideredirect(True)

    window.mainloop()


def load_data_from_file(filename) -> list[dict]:
    dataframe = openpyxl.load_workbook(filename)

    data_list = []
    df = dataframe.active

    header = [col[0].value for col in df.iter_cols(0, df.max_column)]

    print(f'header: {header}')

    for row in range(1, df.max_row):
        row_data = {}
        for i, col in enumerate(df.iter_cols(0, df.max_column)):
            row_data[header[i]] = col[row].value
        data_list.append(row_data)

    return data_list


check_serial()

if __name__ == '__main__':
    t1 = threading.Thread(target=read_usb, daemon=True)
    t1.start()

    time.sleep(1)
    disable_anser_entry()
    time.sleep(1)
    anable_answer_entry()
    time.sleep(1)
    disable_anser_entry()
    time.sleep(1)
    anable_answer_entry()
    time.sleep(1)

    print('end')
    ser.close()
